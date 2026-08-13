import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from pricing_scraper.comparison import (
    ComparisonInputError,
    Size,
    comparison_rows,
    detect_form,
    is_bundle,
    load_own_catalogue,
    match_products,
    name_tokens,
    parse_size,
    prepare,
    score,
    strength_tokens,
    write_comparison,
)
from pricing_scraper.models import Product


def product(
    site: str,
    name: str,
    *,
    brand: str = "Minimalist",
    variant: str = "",
    price: float = 500,
    mrp: float = 600,
    gtin: str = "",
    product_id: str = "",
) -> Product:
    return Product(
        site=site,
        product_id=product_id or f"{site}-{abs(hash(name)) % 10_000}",
        brand=brand,
        product_name=name,
        variant=variant,
        mrp=mrp,
        selling_price=price,
        gtin=gtin,
        product_url=f"https://{site}.test/{abs(hash(name)) % 1000}",
    )


class SizeTests(unittest.TestCase):
    def test_reads_plain_size_and_unit(self):
        self.assertEqual(parse_size("50 ml"), Size(50.0, "ml", 1))
        self.assertEqual(parse_size("Serum (30ml)"), Size(30.0, "ml", 1))
        self.assertEqual(parse_size("1 L Body Wash"), Size(1000.0, "ml", 1))
        self.assertEqual(parse_size("Cream 50 g"), Size(50.0, "g", 1))

    def test_multipack_keeps_unit_size_and_count(self):
        self.assertEqual(parse_size("2 x 50 ml"), Size(50.0, "ml", 2))
        self.assertEqual(parse_size("Cleanser 125ml Pack of 2"), Size(125.0, "ml", 2))

    def test_returns_none_when_no_size_is_stated(self):
        self.assertIsNone(parse_size("Face Serum", ""))

    def test_a_multipack_never_matches_the_single_pack(self):
        single = prepare([product("nykaa", "Gentle Cleanser", variant="125 ml")])[0]
        double = prepare(
            [product("tira", "Gentle Cleanser Pack of 2", variant="125 ml")]
        )[0]
        confidence, method, _ = score(single, double)
        self.assertEqual(confidence, 0.0)
        # The bundle guard fires before the pack-count check; either is correct.
        self.assertIn(method, {"size", "bundle"})


class NormalizationTests(unittest.TestCase):
    def test_tokens_drop_brand_size_and_filler_words(self):
        tokens = name_tokens(
            "Minimalist Niacinamide Face Serum for Oily Skin 30ml", "Minimalist"
        )
        self.assertIn("niacinamide", tokens)
        self.assertIn("oily", tokens)
        self.assertNotIn("minimalist", tokens)
        self.assertNotIn("skin", tokens)
        self.assertNotIn("30", tokens)

    def test_detects_product_form_from_the_longest_keyword(self):
        self.assertEqual(detect_form("Ultra Light Sunscreen SPF 50"), "sunscreen")
        self.assertEqual(detect_form("Hydrating Face Wash"), "cleanser")
        self.assertEqual(detect_form("Vitamin C Face Serum"), "serum")
        self.assertEqual(detect_form("Deep Cleansing Balm"), "cleansing_balm")

    def test_reads_concentrations_and_spf(self):
        self.assertEqual(strength_tokens("10% Niacinamide"), frozenset({"pct10"}))
        self.assertEqual(
            strength_tokens("Sunscreen SPF 50 PA++++"), frozenset({"spf50"})
        )

    def test_recognizes_kits_and_combos(self):
        self.assertTrue(is_bundle("Brightening Combo for Dry Skin"))
        self.assertFalse(is_bundle("Brightening Face Cream"))


class ScoringTests(unittest.TestCase):
    def pair(self, left: Product, right: Product):
        items = prepare([left, right])
        return score(items[0], items[1])

    def test_identical_barcodes_match_outright(self):
        confidence, method, _ = self.pair(
            product("nykaa", "Anything At All", gtin="8901030865275"),
            product("tira", "Different Wording Entirely", gtin="8901030865275"),
        )
        self.assertEqual(method, "gtin")
        self.assertEqual(confidence, 1.0)

    def test_different_barcodes_never_match(self):
        confidence, method, _ = self.pair(
            product("nykaa", "Niacinamide Serum", gtin="8901030865275"),
            product("tira", "Niacinamide Serum", gtin="8904245701192"),
        )
        self.assertEqual(confidence, 0.0)
        self.assertEqual(method, "gtin")

    def test_different_brands_never_match(self):
        confidence, _, _ = self.pair(
            product("nykaa", "Niacinamide Serum 10%"),
            product("tira", "Niacinamide Serum 10%", brand="Dot & Key"),
        )
        self.assertEqual(confidence, 0.0)

    def test_a_sunscreen_never_matches_a_moisturizer(self):
        confidence, method, _ = self.pair(
            product("amazon", "Ceramide Vitamin C Sunscreen", variant="50 g"),
            product("tira", "Ceramide Vitamin C Moisturizing Cream", variant="50 g"),
        )
        self.assertEqual(confidence, 0.0)
        self.assertEqual(method, "form")

    def test_different_concentrations_never_match(self):
        confidence, method, _ = self.pair(
            product("nykaa", "Retinol 0.3% Face Serum", variant="30 ml"),
            product("tira", "Retinol 0.6% Face Serum", variant="30 ml"),
        )
        self.assertEqual(confidence, 0.0)
        self.assertEqual(method, "strength")

    def test_different_sizes_never_match(self):
        confidence, method, _ = self.pair(
            product("nykaa", "Niacinamide Face Serum", variant="30 ml"),
            product("tira", "Niacinamide Face Serum", variant="10 ml"),
        )
        self.assertEqual(confidence, 0.0)
        self.assertEqual(method, "size")

    def test_rounded_sizes_still_match(self):
        confidence, _, _ = self.pair(
            product("nykaa", "Niacinamide Face Serum", variant="50 g"),
            product("tira", "Niacinamide Face Serum", variant="50.0 g"),
        )
        self.assertGreater(confidence, 0.9)

    def test_same_product_with_marketing_tail_matches(self):
        confidence, method, _ = self.pair(
            product(
                "amazon",
                "Minimalist 10% Niacinamide Face Serum for Acne Marks, 30ml",
                variant="30 ml",
            ),
            product("tira", "Minimalist 10% Niacinamide Serum (30 ml)", variant="30 ml"),
        )
        self.assertEqual(method, "name")
        self.assertGreaterEqual(confidence, 0.70)

    def test_missing_size_lowers_confidence(self):
        with_size, _, _ = self.pair(
            product("nykaa", "Niacinamide Face Serum", variant="30 ml"),
            product("tira", "Niacinamide Face Serum", variant="30 ml"),
        )
        without_size, _, notes = self.pair(
            product("nykaa", "Niacinamide Face Serum"),
            product("tira", "Niacinamide Face Serum", variant="30 ml"),
        )
        self.assertLess(without_size, with_size)
        self.assertIn("size missing on one side", notes)


class MatchingTests(unittest.TestCase):
    def test_groups_one_row_per_platform_and_anchors_on_own_catalogue(self):
        report = match_products(
            [
                product("tira", "Niacinamide Face Serum (30 ml)", price=520),
                product("roopsee", "Niacinamide Face Serum", variant="30 ml", price=499),
                product("nykaa", "Niacinamide Face Serum 30ml", price=530),
            ]
        )
        self.assertEqual(len(report.matches), 1)
        match = report.matches[0]
        self.assertEqual(match.anchor.site, "roopsee")
        self.assertEqual(match.sites, ["nykaa", "roopsee", "tira"])
        self.assertEqual(report.unmatched, [])

    def test_a_row_is_never_used_by_two_matches(self):
        report = match_products(
            [
                product("nykaa", "Niacinamide Face Serum", variant="30 ml"),
                product("nykaa", "Niacinamide Face Serum Bright", variant="30 ml"),
                product("tira", "Niacinamide Face Serum", variant="30 ml"),
            ]
        )
        used = [item.product.product_id for match in report.matches
                for item in match.members.values()]
        self.assertEqual(len(used), len(set(used)))

    def test_products_on_one_platform_stay_unmatched(self):
        report = match_products(
            [
                product("nykaa", "Unique Ceramide Cream", variant="50 g"),
                product("tira", "Totally Different Clay Mask", variant="100 g"),
            ]
        )
        self.assertEqual(report.matches, [])
        self.assertEqual(len(report.unmatched), 2)

    def test_rows_carry_a_price_column_per_platform_and_the_cheapest(self):
        report = match_products(
            [
                product("roopsee", "Niacinamide Face Serum", variant="30 ml", price=560),
                product("nykaa", "Niacinamide Face Serum 30ml", price=530),
                product("tira", "Niacinamide Face Serum (30 ml)", price=610),
            ]
        )
        rows = comparison_rows(
            report, ["roopsee", "nykaa", "tira"], own_site="roopsee"
        )
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["roopsee_selling_price"], 560)
        self.assertEqual(row["nykaa_selling_price"], 530)
        self.assertEqual(row["tira_selling_price"], 610)
        self.assertEqual(row["cheapest_platform"], "nykaa")
        self.assertEqual(row["price_gap"], 80.0)
        self.assertEqual(row["roopsee_vs_cheapest"], 30.0)


class OwnCatalogueTests(unittest.TestCase):
    def write_csv(self, directory: Path, rows: list[list[str]]) -> Path:
        path = directory / "own.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            csv.writer(handle).writerows(rows)
        return path

    def test_reads_shopify_style_headers(self):
        with tempfile.TemporaryDirectory() as raw:
            path = self.write_csv(
                Path(raw),
                [
                    ["Handle", "Title", "Vendor", "Option1 Value",
                     "Variant Price", "Variant Compare At Price", "Available"],
                    ["serum-1", "Niacinamide Face Serum", "Minimalist", "30 ml",
                     "499.00", "699.00", "TRUE"],
                ],
            )
            products = load_own_catalogue(path)
        self.assertEqual(len(products), 1)
        loaded = products[0]
        self.assertEqual(loaded.site, "roopsee")
        self.assertEqual(loaded.brand, "Minimalist")
        self.assertEqual(loaded.selling_price, 499.0)
        self.assertEqual(loaded.mrp, 699.0)
        self.assertEqual(loaded.discount_pct, 28.61)
        self.assertIs(loaded.in_stock, True)

    def test_rejects_a_file_without_the_required_columns(self):
        with tempfile.TemporaryDirectory() as raw:
            path = self.write_csv(
                Path(raw), [["item", "cost"], ["Some serum", "499"]]
            )
            with self.assertRaises(ComparisonInputError):
                load_own_catalogue(path)

    def test_rejects_an_unsupported_format(self):
        with tempfile.TemporaryDirectory() as raw:
            path = Path(raw) / "catalogue.txt"
            path.write_text("nope", encoding="utf-8")
            with self.assertRaises(ComparisonInputError):
                load_own_catalogue(path)


class WriteComparisonTests(unittest.TestCase):
    def test_writes_workbook_sheets_and_matching_csv(self):
        report = match_products(
            [
                product("roopsee", "Niacinamide Face Serum", variant="30 ml", price=560),
                product("nykaa", "Niacinamide Face Serum 30ml", price=530),
                product("tira", "Only Here Clay Mask", brand="Deyga", variant="100 g"),
            ]
        )
        with tempfile.TemporaryDirectory() as raw:
            excel_path = Path(raw) / "comparison.xlsx"
            result = write_comparison(report, excel_path)
            self.assertTrue(result.excel_path.exists())
            self.assertTrue(result.csv_path.exists())
            self.assertEqual(result.matches_written, 1)
            self.assertEqual(result.unmatched_rows, 1)

            workbook = load_workbook(result.excel_path)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["comparison", "review", "single_platform"],
                )
                headers = [cell.value for cell in workbook["comparison"][1]]
                self.assertIn("roopsee_selling_price", headers)
                self.assertIn("cheapest_platform", headers)
            finally:
                workbook.close()

            with result.csv_path.open(encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["cheapest_platform"], "nykaa")


if __name__ == "__main__":
    unittest.main()
