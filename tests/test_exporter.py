import csv
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from pricing_scraper.exporter import (
    OUTPUT_COLUMNS,
    deduplicate,
    export_products,
    load_products_csv,
    merge_with_existing_sites,
)
from pricing_scraper.database import DatabaseSyncResult
from pricing_scraper.models import Product


def product(
    product_id: str,
    *,
    site: str = "nykaa",
    price: float = 750,
    scraped_at: str = "2026-07-28T00:00:00+00:00",
) -> Product:
    return Product(
        site=site,
        product_id=product_id,
        brand="Test Brand",
        product_name=f"Product {product_id}",
        variant="50 ml",
        mrp=1000,
        selling_price=price,
        discount_pct=25,
        rating=4.5,
        rating_count=100,
        in_stock=True,
        product_url=f"https://example.test/{product_id}",
        image_url=f"https://images.example.test/{product_id}-1.jpg",
        image_urls=[
            f"https://images.example.test/{product_id}-1.jpg",
            f"https://images.example.test/{product_id}-2.jpg",
        ],
        scraped_at=scraped_at,
    )


class ExporterTests(unittest.TestCase):
    def test_deduplicates_by_site_and_product_id_using_newest_record(self):
        older = product("1", price=800, scraped_at="2026-07-27T00:00:00+00:00")
        newer = product("1", price=700, scraped_at="2026-07-28T00:00:00+00:00")
        older.categories = ["Moisturizers"]
        newer.categories = ["Korean Beauty"]
        other_site = product("1", site="tira")
        result = deduplicate([older, newer, other_site])
        self.assertEqual(len(result), 2)
        nykaa = next(item for item in result if item.site == "nykaa")
        self.assertEqual(nykaa.selling_price, 700)
        self.assertEqual(
            nykaa.categories,
            ["Korean Beauty", "Moisturizers"],
        )

    def test_writes_formatted_excel_and_combined_csv(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            result = export_products(
                [product("1"), product("2")],
                root / "pricing.xlsx",
                root / "pricing.csv",
            )
            self.assertEqual(result.products_written, 2)
            self.assertTrue(result.excel_path.exists())
            self.assertTrue(result.csv_path.exists())

            workbook = load_workbook(result.excel_path)
            self.assertEqual(
                workbook.sheetnames,
                ["combined", "nykaa", "images", "reviews"],
            )
            sheet = workbook["combined"]
            self.assertEqual(
                [cell.value for cell in sheet[1]],
                list(OUTPUT_COLUMNS),
            )
            self.assertEqual(sheet.freeze_panes, "A2")
            mrp_column = OUTPUT_COLUMNS.index("mrp") + 1
            self.assertEqual(
                sheet.cell(2, mrp_column).number_format,
                "₹#,##0.00",
            )
            images = workbook["images"]
            self.assertEqual(images.max_row, 5)
            workbook.close()

            with result.csv_path.open(encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.reader(handle))
            self.assertEqual(rows[0], list(OUTPUT_COLUMNS))
            self.assertEqual(len(rows), 3)

    def test_refreshing_one_site_preserves_other_site_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            first = export_products(
                [product("nykaa-1"), product("tira-1", site="tira")],
                root / "pricing.xlsx",
                root / "pricing.csv",
            )
            merged = merge_with_existing_sites(
                [product("tira-2", site="tira")],
                first.csv_path,
                replacing_site="tira",
            )

            self.assertEqual(
                {(item.site, item.product_id) for item in merged},
                {("nykaa", "nykaa-1"), ("tira", "tira-2")},
            )

    def test_files_are_written_before_the_database_is_synchronized(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            statuses = []
            with patch(
                "pricing_scraper.exporter.sync_products_to_database",
                return_value=DatabaseSyncResult(
                    enabled=True,
                    products_written=1,
                    price_points_written=1,
                ),
            ) as sync:
                result = export_products(
                    [product("1")],
                    root / "pricing.xlsx",
                    root / "pricing.csv",
                    sync_database=True,
                    status_callback=statuses.append,
                )

        sync.assert_called_once()
        self.assertEqual(result.database_products_written, 1)
        self.assertIn("Excel and CSV", statuses[0])
        self.assertIn("Supabase", statuses[1])

    def test_a_failed_required_sync_still_leaves_the_export_on_disk(self):
        """A Supabase outage must not discard a collection that already ran.

        DATABASE_SYNC_REQUIRED still stops the run being reported complete, but
        the workbook and CSV are the durable record of hours of rate-limited
        requests and have to survive a network failure at the very last step.
        """
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            excel_path = root / "pricing.xlsx"
            csv_path = root / "pricing.csv"
            with patch(
                "pricing_scraper.exporter.sync_products_to_database",
                side_effect=ConnectionError("The write operation timed out"),
            ):
                with self.assertRaises(ConnectionError):
                    export_products(
                        [product("1")],
                        excel_path,
                        csv_path,
                        sync_database=True,
                    )

            self.assertTrue(excel_path.exists())
            self.assertTrue(csv_path.exists())
            self.assertEqual(
                [item.product_id for item in load_products_csv(csv_path)],
                ["1"],
            )


class HostedExportTests(unittest.TestCase):
    """A small hosted container cannot afford the workbook, and does not need it."""

    def test_skipping_excel_still_writes_the_csv(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            excel_path = root / "pricing.xlsx"
            csv_path = root / "pricing.csv"

            result = export_products(
                [product("1")], excel_path, csv_path, write_excel=False
            )

            self.assertFalse(excel_path.exists())
            self.assertTrue(csv_path.exists())
            self.assertEqual(
                [item.product_id for item in load_products_csv(csv_path)], ["1"]
            )
            self.assertEqual(result.products_written, 1)

    def test_the_workbook_is_built_by_default(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            excel_path = root / "pricing.xlsx"
            export_products([product("1")], excel_path, root / "pricing.csv")
            self.assertTrue(excel_path.exists())

    def test_no_temporary_files_are_left_behind(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            export_products(
                [product("1")], root / "pricing.xlsx", root / "pricing.csv",
                write_excel=False,
            )
            self.assertEqual([p.name for p in root.glob(".*tmp*")], [])


if __name__ == "__main__":
    unittest.main()
