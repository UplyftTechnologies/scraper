"""Import hand-written products that no retailer scrape found.

Some products never turn up in a sweep: a listing the search misses, a SKU sold
somewhere the scrapers do not cover, or one you want in the comparison before
it is live anywhere. This reads a curated CSV or Excel file and inserts only
the rows that do not already exist, leaving every scraped product alone.

A manual row keeps its own ``site`` value, ``manual`` by default, for two
concrete reasons:

- ``merge_with_existing_sites`` replaces one site's rows wholesale on every
  export, so a manual row filed under ``nykaa`` would be deleted by the next
  Nykaa run.
- ``finalize_retailer_scrape_run`` ages rows of the swept site that the sweep
  did not see, so the same row would be counted missing and go inactive.

Filing manual rows under their own site keeps them out of both. A row may
still name a retailer explicitly, and the import warns when it does.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from pricing_scraper.comparison import (
    ROOPSEE_ALIASES,
    ComparisonInputError,
    _boolean,
    _discount,
    _header_map,
    _number,
    _rows_from_file,
    normalize_text,
)
from pricing_scraper.models import Product, brand_key, normalize_gtin

LOGGER = logging.getLogger("pricing_scraper.manual")

MANUAL_SITE = "manual"
RETAILER_SITES = ("nykaa", "tira", "amazon")

# The manual sheet understands everything the own-catalogue reader does, plus a
# site column so one file can cover several platforms.
MANUAL_ALIASES: dict[str, tuple[str, ...]] = {
    **ROOPSEE_ALIASES,
    "site": ("site", "platform", "retailer", "source", "marketplace"),
}

TEMPLATE_COLUMNS = (
    "site",
    "product_id",
    "brand",
    "product_name",
    "variant",
    "sku",
    "gtin",
    "mrp",
    "selling_price",
    "in_stock",
    "categories",
    "product_url",
    "image_url",
)


class ManualImportError(ValueError):
    """Raised when a manual sheet cannot be interpreted."""


@dataclass(slots=True)
class ManualImportResult:
    """What one manual import did."""

    path: Path
    rows_read: int = 0
    inserted: list[Product] = field(default_factory=list)
    skipped_existing: list[tuple[Product, str]] = field(default_factory=list)
    rejected: list[tuple[int, str]] = field(default_factory=list)
    retailer_sites: set[str] = field(default_factory=set)
    written: bool = False
    database_synced: bool = False
    inserted_rows: int = 0

    def summary(self) -> str:
        return (
            f"{self.path.name}: {len(self.inserted):,} new product(s), "
            f"{len(self.skipped_existing):,} already in Supabase, "
            f"{len(self.rejected):,} unusable row(s)"
        )


def _rows_from_upload(payload: bytes, filename: str) -> list[dict[str, Any]]:
    """Read an uploaded CSV or Excel file straight from memory."""
    suffix = Path(filename or "").suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        import io

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            try:
                headers = [str(cell or "") for cell in next(values)]
            except StopIteration:
                return []
            return [
                dict(zip(headers, row))
                for row in values
                if any(cell is not None for cell in row)
            ]
        finally:
            workbook.close()
    if suffix in {".csv", ""}:
        import csv as _csv
        import io

        text = payload.decode("utf-8-sig", errors="replace")
        return [dict(row) for row in _csv.DictReader(io.StringIO(text))]
    raise ManualImportError(
        f"Unsupported file type {suffix!r}; upload a .csv or .xlsx file."
    )


def _identity(site: str, brand: str, name: str) -> tuple[str, str, str]:
    """The brand+name fallback key, folded the way the scrape filter folds."""
    return (site.casefold(), brand_key(brand), normalize_text(name))


def existing_index(
    known: Mapping[str, Mapping[str, Any]] | Iterable[Product],
    *,
    site: str,
) -> tuple[set[str], set[tuple[str, str, str]]]:
    """Index one site's stored products by ID and by brand+name."""
    ids: set[str] = set()
    identities: set[tuple[str, str, str]] = set()
    rows: Iterable[Any]
    if isinstance(known, Mapping):
        rows = known.values()
    else:
        rows = known
    for row in rows:
        if isinstance(row, Product):
            product_id = row.product_id
            brand, name = row.brand, row.product_name
        else:
            product_id = str(row.get("product_id") or "")
            brand = str(row.get("brand") or "")
            name = str(row.get("product_name") or "")
        if product_id:
            ids.add(product_id)
        if brand or name:
            identities.add(_identity(site, brand, name))
    return ids, identities


def load_manual_products(
    source: Path | bytes,
    *,
    default_site: str = MANUAL_SITE,
    filename: str = "",
) -> tuple[list[Product], list[tuple[int, str]]]:
    """Read a sheet into products, reporting the rows it could not use.

    Accepts a path or the raw bytes of an upload, so the dashboard can hand
    over a file the browser sent without writing it to disk first.
    """
    if isinstance(source, (bytes, bytearray)):
        rows = _rows_from_upload(bytes(source), filename)
        label = filename or "upload"
    else:
        resolved = Path(source).expanduser()
        if not resolved.exists():
            raise ManualImportError(f"Manual product file not found: {resolved}")
        try:
            rows = _rows_from_file(resolved)
        except ComparisonInputError as exc:
            raise ManualImportError(str(exc)) from exc
        label = resolved.name
    if not rows:
        raise ManualImportError(f"{label} has no rows.")

    columns = _header_map(rows[0].keys(), MANUAL_ALIASES)
    missing = [
        name for name in ("brand", "product_name") if name not in columns
    ]
    if missing:
        raise ManualImportError(
            f"{label} has no column for {', '.join(missing)}. "
            f"Recognized headers: {', '.join(sorted(columns)) or 'none'}."
        )

    products: list[Product] = []
    rejected: list[tuple[int, str]] = []
    for index, row in enumerate(rows, start=2):  # row 1 is the header
        def cell(name: str) -> Any:
            column = columns.get(name)
            return row.get(column) if column else None

        brand = str(cell("brand") or "").strip()
        name = str(cell("product_name") or "").strip()
        if not name:
            rejected.append((index, "no product name"))
            continue
        if not brand:
            rejected.append((index, f"no brand for {name!r}"))
            continue

        site = str(cell("site") or "").strip().casefold() or default_site
        mrp = _number(cell("mrp"))
        selling_price = _number(cell("selling_price"))
        categories = str(cell("categories") or "").strip()
        product_id = str(cell("product_id") or "").strip()
        sku = str(cell("sku") or "").strip()
        if not product_id:
            # A stable identity so re-importing the same sheet cannot create a
            # second copy of the same product.
            product_id = sku or f"manual-{brand_key(brand)}-{normalize_text(name).replace(' ', '-')}"[:120]
        products.append(
            Product(
                site=site,
                product_id=product_id,
                brand=brand,
                product_name=name,
                categories=[categories] if categories else [],
                sku=sku,
                gtin=normalize_gtin(cell("gtin")),
                variant=str(cell("variant") or "").strip(),
                mrp=mrp,
                selling_price=selling_price,
                discount_pct=_discount(mrp, selling_price),
                in_stock=_boolean(cell("in_stock")),
                product_url=str(cell("product_url") or "").strip(),
                image_url=str(cell("image_url") or "").strip(),
            )
        )
    return products, rejected


def select_new(
    products: Sequence[Product],
    *,
    store: Any,
) -> tuple[list[Product], list[tuple[Product, str]]]:
    """Split the sheet into rows to insert and rows Supabase already holds.

    A product counts as present when its ``site`` and ``product_id`` match a
    stored row, or when its brand and name match one. The second check means a
    hand-written row does not have to carry an ID the retailer assigned.

    Only Supabase is consulted. The local CSV export is deliberately ignored:
    a product sitting in the file but missing from the database is exactly the
    row this is meant to insert.
    """
    by_site: dict[str, list[Product]] = {}
    for product in products:
        by_site.setdefault(product.site, []).append(product)

    fresh: list[Product] = []
    existing: list[tuple[Product, str]] = []
    for site, site_products in by_site.items():
        rows = store.fetch_site_products(
            site, columns=("site", "product_id", "brand", "product_name")
        )
        ids, identities = existing_index(
            {str(row.get("product_id")): row for row in rows}, site=site
        )
        LOGGER.info("manual_check site=%s already_in_supabase=%s", site, len(ids))
        # Rows inside the sheet itself must not collide either.
        seen_ids: set[str] = set()
        seen_identities: set[tuple[str, str, str]] = set()
        for product in site_products:
            identity = _identity(site, product.brand, product.product_name)
            if product.product_id in ids:
                existing.append((product, "product_id already stored"))
            elif identity in identities:
                existing.append((product, "brand and name already stored"))
            elif product.product_id in seen_ids:
                existing.append((product, "duplicate product_id in the sheet"))
            elif identity in seen_identities:
                existing.append((product, "duplicate brand and name in the sheet"))
            else:
                seen_ids.add(product.product_id)
                seen_identities.add(identity)
                fresh.append(product)
    return fresh, existing


def insert_manual_products(
    source: Path | bytes,
    *,
    filename: str = "",
    default_site: str = MANUAL_SITE,
    dry_run: bool = False,
    store: Any = None,
) -> ManualImportResult:
    """Insert the sheet's rows that Supabase does not already hold.

    Nothing is updated and nothing is deleted: a product the database already
    has is reported and left exactly as it is, so running the same sheet twice
    is harmless.
    """
    from pricing_scraper.database import (
        DatabaseConfigurationError,
        SupabaseCatalogStore,
    )

    if store is None:
        store, _required = SupabaseCatalogStore.from_environment()
        if store is None:
            raise DatabaseConfigurationError(
                "Inserting needs Supabase. Set SUPABASE_URL and a server-side "
                "key in .env."
            )

    products, rejected = load_manual_products(
        source, default_site=default_site, filename=filename
    )
    result = ManualImportResult(
        path=Path(filename or (source if isinstance(source, Path) else "upload")),
        rows_read=len(products) + len(rejected),
        rejected=rejected,
    )
    fresh, existing = select_new(products, store=store)
    result.inserted = fresh
    result.skipped_existing = existing
    result.retailer_sites = {
        product.site for product in fresh if product.site in RETAILER_SITES
    }

    if dry_run or not fresh:
        return result

    rows = []
    for product in fresh:
        row = product.to_dict()
        row["scraped_at"] = product.scraped_at or _now()
        row["first_seen_at"] = row["scraped_at"]
        row["last_seen_at"] = row["scraped_at"]
        row["last_checked_at"] = row["scraped_at"]
        row["is_active"] = True
        rows.append(row)
    written = store._upsert(store.products_table, rows, "site,product_id")
    result.written = True
    result.database_synced = True
    result.inserted_rows = written
    LOGGER.info("%s", result.summary())
    return result


def _now() -> str:
    from datetime import datetime, timezone

    return datetime.now(timezone.utc).isoformat(timespec="microseconds")
