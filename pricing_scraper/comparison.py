"""Cross-platform product matching for price comparison.

No retailer identifier is shared across platforms: only Nykaa publishes a
usable GTIN, Tira exposes its internal item code, and Amazon India omits
barcodes on beauty pages. Products are therefore matched on what every
platform does publish - brand, pack size, and the product name - with the
barcode used as an exact shortcut whenever two rows happen to carry one.

Matching is deliberately conservative: a pair must agree on brand, on pack
size, and on product form (a sunscreen never matches a moisturiser), because a
wrong pairing silently misreports a competitor's price.
"""

from __future__ import annotations

import csv
import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from pricing_scraper.models import Product, brand_key, normalize_gtin

LOGGER = logging.getLogger(__name__)

ROOPSEE_SITE = "roopsee"
# Own catalogue first: it anchors the comparison. Nykaa follows because its
# rows carry barcodes, which give the strongest matches.
ANCHOR_ORDER = (ROOPSEE_SITE, "nykaa", "tira", "amazon")

# Words that carry no distinguishing meaning in a beauty product title.
STOP_WORDS = frozenset(
    {
        "a", "all", "an", "and", "at", "best", "buy", "care", "combo", "for",
        "free", "from", "get", "gift", "in", "india", "kit", "men", "new",
        "of", "offer", "online", "pack", "price", "product", "set", "skin",
        "the", "to", "type", "types", "value", "with", "women", "your",
    }
)

# A product's form. Two rows that declare different forms are never the same
# product, however similar their names read.
PRODUCT_FORMS: dict[str, tuple[str, ...]] = {
    "sunscreen": ("sunscreen", "sunblock", "sun screen", "spf", "uv shield"),
    "serum": ("serum", "ampoule", "essence", "booster"),
    "moisturizer": (
        "moisturizer", "moisturiser", "cream", "lotion", "emulsion", "gel cream",
    ),
    "cleanser": (
        "cleanser", "face wash", "facewash", "foam", "cleansing gel",
        "cleansing water", "micellar", "shampoo", "body wash", "gel moussant",
    ),
    "cleansing_balm": ("cleansing balm", "cleansing oil", "makeup remover"),
    "toner": ("toner", "mist", "tonic"),
    "mask": ("mask", "sheet mask", "pack", "peel off"),
    "exfoliant": ("scrub", "exfoliant", "exfoliator", "peel", "peeling"),
    "oil": ("face oil", "body oil", "hair oil", "oil"),
    "balm": ("balm", "salve", "ointment"),
    "lip": ("lip balm", "lipstick", "lip mask", "lip oil", "lip gloss"),
    "eye": ("eye cream", "eye serum", "eye gel", "under eye"),
    "powder": ("powder", "sachet"),
    "soap": ("soap", "bar", "cleansing bar"),
    "supplement": ("tablet", "capsule", "gummies", "sachets", "supplement"),
    "tool": ("brush", "roller", "device", "massager", "tool", "sponge"),
    "deodorant": ("deodorant", "deo", "antiperspirant"),
}
# Forms that retailers use interchangeably for the same product.
FORM_ALIASES = {"balm": "moisturizer", "powder": ""}

# Concentration and SPF are part of a skincare product's identity: 10%
# niacinamide and 5% niacinamide are different products at different prices.
PERCENT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*%")
SPF_RE = re.compile(r"spf\s*(\d+)", re.IGNORECASE)
BUNDLE_RE = re.compile(
    r"\b(combo|bundle|duo|trio|kit|regimen|routine|hamper|set of)\b", re.IGNORECASE
)

UNIT_TO_ML = {"ml": 1.0, "l": 1000.0, "litre": 1000.0, "liter": 1000.0}
UNIT_TO_G = {"g": 1.0, "gm": 1.0, "gms": 1.0, "gram": 1.0, "grams": 1.0, "kg": 1000.0}
COUNT_UNITS = (
    "capsules", "capsule", "tablets", "tablet", "sachets", "sachet",
    "sheets", "sheet", "wipes", "pcs", "piece", "pieces", "count",
)
SIZE_RE = re.compile(
    r"(\d+(?:[.,]\d+)?)\s*(ml|l|litre|liter|g|gm|gms|gram|grams|kg|oz|"
    + "|".join(COUNT_UNITS)
    + r")\b",
    re.IGNORECASE,
)
MULTI_RE = re.compile(
    r"(\d+)\s*[x×*]\s*(\d+(?:[.,]\d+)?)\s*(ml|l|g|gm|gms|kg)\b", re.IGNORECASE
)
PACK_RE = re.compile(
    r"(?:pack|set|combo|box)\s*of\s*(\d+)|(\d+)\s*[- ]?(?:pack|piece|pc)s?\b",
    re.IGNORECASE,
)


@dataclass(frozen=True, slots=True)
class Size:
    """A normalized pack size: unit amount, its unit, and how many units."""

    amount: float
    unit: str
    pack: int = 1

    @property
    def total(self) -> float:
        return self.amount * self.pack

    def label(self) -> str:
        base = f"{self.amount:g}{self.unit}"
        return f"{self.pack}x{base}" if self.pack > 1 else base


@dataclass(slots=True)
class Item:
    """One platform's row, prepared for matching."""

    product: Product
    site: str
    bkey: str
    tokens: frozenset[str]
    size: Size | None
    form: str
    gtin: str
    strengths: frozenset[str] = frozenset()
    bundle: bool = False

    @property
    def name(self) -> str:
        return self.product.product_name


@dataclass(slots=True)
class Match:
    """One matched product across two or more platforms."""

    anchor: Item
    members: dict[str, Item] = field(default_factory=dict)
    confidence: float = 1.0
    method: str = "name"
    notes: list[str] = field(default_factory=list)

    @property
    def sites(self) -> list[str]:
        return sorted(self.members)


@dataclass(frozen=True, slots=True)
class MatchReport:
    """Everything one matching pass produced."""

    matches: list[Match]
    unmatched: list[Item]
    items_by_site: dict[str, int]
    threshold: float


def normalize_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).strip()


def parse_size(*texts: str) -> Size | None:
    """Read a pack size out of a variant label or product title.

    ``2 x 50 ml`` and ``Pack of 2 (50ml)`` both describe two 50 ml units, which
    must not be confused with a single 100 ml bottle.
    """
    joined = " ".join(str(text or "") for text in texts)
    if not joined.strip():
        return None

    multi = MULTI_RE.search(joined)
    if multi:
        pack = int(multi.group(1))
        amount, unit = _canonical_amount(multi.group(2), multi.group(3))
        if unit:
            return Size(amount, unit, max(1, pack))

    pack = 1
    pack_match = PACK_RE.search(joined)
    if pack_match:
        pack = int(pack_match.group(1) or pack_match.group(2) or 1)

    for raw_amount, raw_unit in SIZE_RE.findall(joined):
        amount, unit = _canonical_amount(raw_amount, raw_unit)
        if unit:
            return Size(amount, unit, max(1, pack))
    return None


def _canonical_amount(raw_amount: str, raw_unit: str) -> tuple[float, str]:
    try:
        amount = float(str(raw_amount).replace(",", "."))
    except ValueError:
        return 0.0, ""
    unit = str(raw_unit).casefold()
    if unit in UNIT_TO_ML:
        return amount * UNIT_TO_ML[unit], "ml"
    if unit in UNIT_TO_G:
        return amount * UNIT_TO_G[unit], "g"
    if unit == "oz":
        return round(amount * 29.5735, 1), "ml"
    if unit in COUNT_UNITS:
        return amount, "pc"
    return 0.0, ""


def detect_form(*texts: str) -> str:
    """Classify the product form, preferring the most specific keyword."""
    haystack = " ".join(normalize_text(text) for text in texts)
    best = ""
    best_length = 0
    for form, keywords in PRODUCT_FORMS.items():
        for keyword in keywords:
            if keyword in haystack and len(keyword) > best_length:
                best, best_length = form, len(keyword)
    return FORM_ALIASES.get(best, best)


def strength_tokens(*texts: str) -> frozenset[str]:
    """Active-ingredient percentages and SPF values stated in a title."""
    joined = " ".join(str(text or "") for text in texts)
    found = {f"pct{float(value):g}" for value in PERCENT_RE.findall(joined)}
    found.update(f"spf{int(value)}" for value in SPF_RE.findall(joined))
    return frozenset(found)


def is_bundle(*texts: str) -> bool:
    """True when the title sells a multi-product kit rather than one item."""
    return bool(BUNDLE_RE.search(" ".join(str(text or "") for text in texts)))


def name_tokens(name: str, brand: str) -> frozenset[str]:
    """Distinguishing words of a title, without the brand or the pack size."""
    text = MULTI_RE.sub(" ", str(name or ""))
    text = SIZE_RE.sub(" ", text)
    text = PACK_RE.sub(" ", text)
    words = normalize_text(text).split()
    brand_words = set(normalize_text(brand).split())
    return frozenset(
        word
        for word in words
        if word not in STOP_WORDS and word not in brand_words and not word.isdigit()
    )


def prepare(products: Iterable[Product]) -> list[Item]:
    """Wrap products with the keys matching needs."""
    items = []
    for product in products:
        site = str(product.site or "").casefold().strip()
        if not site:
            continue
        tokens = name_tokens(product.product_name, product.brand)
        items.append(
            Item(
                product=product,
                site=site,
                bkey=brand_key(product.brand),
                tokens=tokens,
                size=parse_size(product.variant, product.product_name),
                form=detect_form(product.product_name, product.variant),
                gtin=normalize_gtin(product.gtin),
                strengths=strength_tokens(product.product_name, product.variant),
                bundle=is_bundle(product.product_name, product.variant),
            )
        )
    return items


def score(left: Item, right: Item) -> tuple[float, str, list[str]]:
    """Rate how likely two rows are the same product.

    Returns the confidence, the method that produced it, and any caveats worth
    recording in the review sheet.
    """
    if left.bkey != right.bkey or not left.bkey:
        return 0.0, "brand", ["different brand"]
    if left.gtin and right.gtin:
        if left.gtin == right.gtin:
            return 1.0, "gtin", []
        return 0.0, "gtin", ["different barcode"]

    if not left.tokens or not right.tokens:
        return 0.0, "name", ["no comparable words in the title"]
    shared = left.tokens & right.tokens
    if len(shared) < 2:
        return 0.0, "name", ["fewer than two shared words"]

    if left.form and right.form and left.form != right.form:
        return 0.0, "form", [f"{left.form} vs {right.form}"]
    if left.strengths and right.strengths and left.strengths != right.strengths:
        return 0.0, "strength", [
            f"{'/'.join(sorted(left.strengths))} vs "
            f"{'/'.join(sorted(right.strengths))}"
        ]
    if left.bundle != right.bundle:
        # A combo/kit listing is a different item from the single product.
        return 0.0, "bundle", ["one side is a kit or combo"]

    containment = len(shared) / min(len(left.tokens), len(right.tokens))
    jaccard = len(shared) / len(left.tokens | right.tokens)
    confidence = 0.65 * containment + 0.35 * jaccard
    notes: list[str] = []

    if left.size and right.size:
        if left.size.pack != right.size.pack:
            return 0.0, "size", [
                f"pack of {left.size.pack} vs {right.size.pack}"
            ]
        if left.size.unit != right.size.unit or not _amounts_agree(
            left.size.amount, right.size.amount
        ):
            return 0.0, "size", [
                f"{left.size.label()} vs {right.size.label()}"
            ]
        confidence += 0.10
    else:
        # Without a size on both sides the pairing rests on wording alone.
        confidence *= 0.85
        notes.append("size missing on one side")

    if left.form and right.form:
        confidence += 0.05
    elif not left.form and not right.form:
        notes.append("product form not recognized")

    return min(1.0, confidence), "name", notes


def _amounts_agree(left: float, right: float) -> bool:
    """Allow the rounding retailers apply to the same pack (50 g vs 50.0 g)."""
    if left <= 0 or right <= 0:
        return False
    return abs(left - right) <= max(0.5, 0.02 * max(left, right))


def match_products(
    products: Iterable[Product],
    *,
    threshold: float = 0.70,
    anchor_order: Sequence[str] = ANCHOR_ORDER,
) -> MatchReport:
    """Group rows from different platforms into one product per match.

    Each platform contributes at most one row per match: the best-scoring one.
    Anchors are taken in ``anchor_order`` so the comparison is organized around
    the catalogue that matters most, and every row is consumed at most once.
    """
    items = prepare(products)
    by_site: dict[str, list[Item]] = {}
    for item in items:
        by_site.setdefault(item.site, []).append(item)

    # Bucket by brand so scoring never runs across the whole catalogue.
    buckets: dict[str, list[Item]] = {}
    for item in items:
        buckets.setdefault(item.bkey, []).append(item)

    consumed: set[int] = set()
    matches: list[Match] = []
    ordered_sites = [site for site in anchor_order if site in by_site]
    ordered_sites += sorted(site for site in by_site if site not in anchor_order)

    for anchor_site in ordered_sites:
        for anchor in by_site[anchor_site]:
            if id(anchor) in consumed or not anchor.bkey:
                continue
            match = Match(anchor=anchor, members={anchor_site: anchor})
            confidences: list[float] = []
            methods: set[str] = set()

            available: dict[str, list[Item]] = {}
            for option in buckets.get(anchor.bkey, ()):
                if option.site == anchor_site or id(option) in consumed:
                    continue
                available.setdefault(option.site, []).append(option)

            for site, options in sorted(available.items()):
                best: Item | None = None
                best_confidence = 0.0
                best_method = ""
                best_notes: list[str] = []
                for option in options:
                    confidence, method, notes = score(anchor, option)
                    if confidence >= threshold and confidence > best_confidence:
                        best, best_confidence = option, confidence
                        best_method, best_notes = method, notes
                if best is not None:
                    match.members[site] = best
                    consumed.add(id(best))
                    confidences.append(best_confidence)
                    methods.add(best_method)
                    match.notes.extend(f"{site}: {note}" for note in best_notes)
            if len(match.members) > 1:
                consumed.add(id(anchor))
                match.confidence = min(confidences)
                match.method = "gtin" if methods == {"gtin"} else "name"
                matches.append(match)

    unmatched = [item for item in items if id(item) not in consumed]
    matches.sort(
        key=lambda item: (
            item.anchor.product.brand.casefold(),
            item.anchor.name.casefold(),
        )
    )
    return MatchReport(
        matches=matches,
        unmatched=unmatched,
        items_by_site={site: len(rows) for site, rows in sorted(by_site.items())},
        threshold=threshold,
    )


# --------------------------------------------------------------------------
# Loading the platforms
# --------------------------------------------------------------------------

# Header spellings seen in storefront and marketplace exports.
ROOPSEE_ALIASES: dict[str, tuple[str, ...]] = {
    "product_id": ("product_id", "id", "handle", "product handle", "item id"),
    "sku": ("sku", "variant sku", "sku code", "item code"),
    "gtin": ("gtin", "ean", "upc", "barcode", "variant barcode"),
    "brand": ("brand", "vendor", "brand name", "manufacturer"),
    "product_name": ("product_name", "title", "name", "product title", "product"),
    "variant": (
        "variant", "variant title", "size", "weight", "pack size", "option1 value",
    ),
    "mrp": (
        "mrp", "compare at price", "compare_at_price", "list price", "variant compare at price",
    ),
    "selling_price": (
        "selling_price", "price", "sale price", "variant price", "offer price",
        "discounted price",
    ),
    "in_stock": ("in_stock", "available", "stock", "inventory", "status"),
    "product_url": ("product_url", "url", "link", "product link"),
    "image_url": ("image_url", "image", "image src", "images"),
    "categories": ("categories", "category", "collection", "type", "product type"),
}


# The only stored columns matching and the comparison sheet actually read.
MATCH_COLUMNS = (
    "site",
    "product_id",
    "parent_product_id",
    "sku",
    "gtin",
    "brand",
    "product_name",
    "variant",
    "categories",
    "mrp",
    "selling_price",
    "discount_pct",
    "in_stock",
    "product_url",
)


class ComparisonInputError(ValueError):
    """Raised when a supplied catalogue file cannot be interpreted."""


def _header_map(
    headers: Iterable[str],
    alias_table: Mapping[str, Sequence[str]] | None = None,
) -> dict[str, str]:
    """Map our field names onto whatever the export happens to call them."""
    normalized = {
        normalize_text(header).replace(" ", "_"): header
        for header in headers
        if str(header or "").strip()
    }
    resolved: dict[str, str] = {}
    for field_name, aliases in (alias_table or ROOPSEE_ALIASES).items():
        for alias in aliases:
            key = normalize_text(alias).replace(" ", "_")
            if key in normalized:
                resolved[field_name] = normalized[key]
                break
    return resolved


def _number(value: Any) -> float | None:
    text = re.sub(r"[^\d.\-]", "", str(value or "").replace(",", ""))
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if number >= 0 else None


def _boolean(value: Any) -> bool | None:
    folded = str(value or "").strip().casefold()
    if folded in {"true", "yes", "y", "1", "in stock", "instock", "active", "available"}:
        return True
    if folded in {"false", "no", "n", "0", "out of stock", "sold out", "draft"}:
        return False
    return None


def _rows_from_file(path: Path) -> list[dict[str, Any]]:
    """Read a CSV or Excel export into plain dictionaries."""
    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            try:
                headers = [str(cell or "") for cell in next(rows)]
            except StopIteration:
                return []
            return [
                dict(zip(headers, values))
                for values in rows
                if any(value is not None for value in values)
            ]
        finally:
            workbook.close()
    if path.suffix.casefold() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    raise ComparisonInputError(
        f"Unsupported catalogue format {path.suffix!r}; use .csv or .xlsx."
    )


def load_own_catalogue(
    path: Path, *, site: str = ROOPSEE_SITE
) -> list[Product]:
    """Load an own-catalogue export, mapping its headers onto Product fields."""
    resolved = path.expanduser()
    if not resolved.exists():
        raise ComparisonInputError(f"Catalogue file not found: {resolved}")
    rows = _rows_from_file(resolved)
    if not rows:
        raise ComparisonInputError(f"Catalogue file is empty: {resolved}")

    columns = _header_map(rows[0].keys())
    missing = [
        field_name
        for field_name in ("brand", "product_name", "selling_price")
        if field_name not in columns
    ]
    if missing:
        raise ComparisonInputError(
            f"{resolved.name} has no column for {', '.join(missing)}. "
            f"Recognized headers: {', '.join(sorted(columns)) or 'none'}. "
            "Rename the columns or add an alias to ROOPSEE_ALIASES."
        )
    LOGGER.info(
        "%s columns mapped: %s",
        resolved.name,
        ", ".join(f"{ours}<-{theirs}" for ours, theirs in sorted(columns.items())),
    )

    products: list[Product] = []
    for index, row in enumerate(rows, start=1):
        def cell(field_name: str) -> Any:
            column = columns.get(field_name)
            return row.get(column) if column else None

        name = str(cell("product_name") or "").strip()
        if not name:
            continue
        mrp = _number(cell("mrp"))
        selling_price = _number(cell("selling_price"))
        categories = str(cell("categories") or "").strip()
        products.append(
            Product(
                site=site,
                product_id=str(cell("product_id") or cell("sku") or index).strip(),
                brand=str(cell("brand") or "").strip(),
                product_name=name,
                categories=[categories] if categories else [],
                sku=str(cell("sku") or "").strip(),
                gtin=normalize_gtin(cell("gtin")),
                variant=str(cell("variant") or "").strip(),
                mrp=mrp,
                selling_price=selling_price,
                discount_pct=_discount(mrp, selling_price),
                in_stock=_boolean(cell("in_stock")),
                product_url=str(cell("product_url") or "").strip(),
                image_url=str(cell("image_url") or "").strip(),
            )
        )
    return products


def _discount(mrp: float | None, selling_price: float | None) -> float | None:
    if not mrp or selling_price is None or mrp <= 0 or selling_price > mrp:
        return None
    return round((mrp - selling_price) / mrp * 100, 2)


def load_retailer_products(
    *,
    csv_path: Path | None = None,
    use_database: bool = True,
    sites: Sequence[str] = ("nykaa", "tira", "amazon"),
) -> list[Product]:
    """Read scraped retailer rows from Supabase, falling back to the CSV export."""
    if use_database:
        store = None
        try:
            from pricing_scraper.database import SupabaseCatalogStore

            # from_environment returns (store | None, sync_required); a missing
            # URL and key simply means the database is not configured here.
            store, _ = SupabaseCatalogStore.from_environment()
        except Exception as exc:
            LOGGER.info("Database unavailable (%s); using the CSV export.", exc)
        if store is not None:
            products: list[Product] = []
            try:
                for site in sites:
                    # Matching never looks at descriptions, galleries or
                    # reviews, and leaving them out keeps the read fast.
                    rows = store.fetch_site_products(site, columns=MATCH_COLUMNS)
                    products.extend(_product_from_row(row) for row in rows)
                    LOGGER.info("%s: %s rows from the database", site, len(rows))
            except Exception as exc:
                LOGGER.warning(
                    "Database read failed (%s); using the CSV export.", exc
                )
                products = []
            if products:
                return products
            LOGGER.info("No database rows available; using the CSV export.")

    if csv_path is None:
        raise ComparisonInputError(
            "No retailer data: the database is unavailable and no CSV export "
            "path was given."
        )
    from pricing_scraper.exporter import load_products_csv

    products = load_products_csv(Path(csv_path))
    if not products:
        raise ComparisonInputError(f"No retailer rows found in {csv_path}.")
    return products


PLATFORM_COLUMNS = (
    "selling_price",
    "mrp",
    "discount_pct",
    "in_stock",
    "name",
    "url",
)
BASE_COLUMNS = (
    "brand",
    "product",
    "form",
    "size",
    "pack",
    "categories",
    "platforms",
    "platform_count",
    "match_method",
    "confidence",
    "notes",
)
SUMMARY_COLUMNS = (
    "min_price",
    "max_price",
    "price_gap",
    "price_gap_pct",
    "cheapest_platform",
)
UNMATCHED_COLUMNS = (
    "site",
    "brand",
    "product_name",
    "variant",
    "size",
    "form",
    "mrp",
    "selling_price",
    "in_stock",
    "product_url",
)


def comparison_columns(platforms: Sequence[str], *, own_site: str = "") -> list[str]:
    """Header order: identity, one block per platform, then the comparison."""
    columns = list(BASE_COLUMNS)
    for platform in platforms:
        columns.extend(f"{platform}_{field}" for field in PLATFORM_COLUMNS)
    columns.extend(SUMMARY_COLUMNS)
    if own_site and own_site in platforms:
        columns.extend((f"{own_site}_vs_cheapest", f"{own_site}_vs_cheapest_pct"))
    return columns


def comparison_rows(
    report: MatchReport,
    platforms: Sequence[str],
    *,
    own_site: str = "",
) -> list[dict[str, Any]]:
    """One dictionary per matched product, with a column block per platform."""
    rows = []
    for match in report.matches:
        anchor = match.anchor
        row: dict[str, Any] = {
            "brand": anchor.product.brand,
            "product": anchor.name,
            "form": anchor.form,
            "size": anchor.size.label() if anchor.size else "",
            "pack": anchor.size.pack if anchor.size else 1,
            "categories": ", ".join(anchor.product.categories),
            "platforms": ", ".join(match.sites),
            "platform_count": len(match.members),
            "match_method": match.method,
            "confidence": round(match.confidence, 3),
            "notes": "; ".join(match.notes),
        }
        prices: dict[str, float] = {}
        for platform in platforms:
            item = match.members.get(platform)
            product = item.product if item else None
            price = product.selling_price if product else None
            if price is not None:
                prices[platform] = float(price)
            row[f"{platform}_selling_price"] = price
            row[f"{platform}_mrp"] = product.mrp if product else None
            row[f"{platform}_discount_pct"] = product.discount_pct if product else None
            row[f"{platform}_in_stock"] = product.in_stock if product else None
            row[f"{platform}_name"] = product.product_name if product else ""
            row[f"{platform}_url"] = product.product_url if product else ""

        if prices:
            cheapest = min(prices, key=lambda name: prices[name])
            low, high = prices[cheapest], max(prices.values())
            row["min_price"] = low
            row["max_price"] = high
            row["price_gap"] = round(high - low, 2)
            row["price_gap_pct"] = round((high - low) / low * 100, 2) if low else None
            row["cheapest_platform"] = cheapest
            if own_site and own_site in platforms:
                own = prices.get(own_site)
                others = {
                    name: value for name, value in prices.items() if name != own_site
                }
                if own is not None and others:
                    rival = min(others.values())
                    row[f"{own_site}_vs_cheapest"] = round(own - rival, 2)
                    row[f"{own_site}_vs_cheapest_pct"] = (
                        round((own - rival) / rival * 100, 2) if rival else None
                    )
        rows.append(row)
    return rows


def unmatched_rows(report: MatchReport) -> list[dict[str, Any]]:
    """Products found on a single platform, for coverage and gap analysis."""
    rows = [
        {
            "site": item.site,
            "brand": item.product.brand,
            "product_name": item.name,
            "variant": item.product.variant,
            "size": item.size.label() if item.size else "",
            "form": item.form,
            "mrp": item.product.mrp,
            "selling_price": item.product.selling_price,
            "in_stock": item.product.in_stock,
            "product_url": item.product.product_url,
        }
        for item in report.unmatched
    ]
    rows.sort(
        key=lambda row: (
            row["site"],
            str(row["brand"]).casefold(),
            str(row["product_name"]).casefold(),
        )
    )
    return rows


@dataclass(frozen=True, slots=True)
class ComparisonResult:
    """Paths and counts produced by one comparison export."""

    excel_path: Path
    csv_path: Path
    matches_written: int
    review_rows: int
    unmatched_rows: int
    platforms: tuple[str, ...]
    matches_by_platforms: dict[str, int]


def write_comparison(
    report: MatchReport,
    excel_path: Path,
    csv_path: Path | None = None,
    *,
    own_site: str = ROOPSEE_SITE,
    review_below: float = 0.80,
) -> ComparisonResult:
    """Write the comparison workbook and its flat CSV counterpart."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    platforms = [site for site in ANCHOR_ORDER if site in report.items_by_site]
    platforms += sorted(
        site for site in report.items_by_site if site not in ANCHOR_ORDER
    )
    columns = comparison_columns(platforms, own_site=own_site)
    rows = comparison_rows(report, platforms, own_site=own_site)
    review = [
        row
        for row in rows
        if float(row["confidence"]) < review_below or row["notes"]
    ]
    gaps = unmatched_rows(report)

    excel_path = excel_path.resolve()
    csv_path = (
        csv_path.resolve()
        if csv_path
        else excel_path.with_name(f"{excel_path.stem}.csv")
    )
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    def sheet_for(workbook: Workbook, title: str, headers: Sequence[str], data):
        sheet = workbook.create_sheet(title=title[:31])
        sheet.append(list(headers))
        for row in data:
            sheet.append([row.get(header) for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        fill = PatternFill("solid", fgColor="D81B60")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
        for index, header in enumerate(headers, start=1):
            if header.endswith(("price", "mrp", "price_gap", "vs_cheapest")):
                number_format = "₹#,##0.00"
            elif header.endswith(("pct", "confidence")):
                number_format = "0.00"
            else:
                continue
            for cell in sheet.iter_rows(
                min_row=2, min_col=index, max_col=index
            ):
                cell[0].number_format = number_format
        for column_cells in sheet.columns:
            width = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in column_cells
            )
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(width + 2, 10), 55
            )
        return sheet

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_for(workbook, "comparison", columns, rows)
    sheet_for(workbook, "review", columns, review)
    sheet_for(workbook, "single_platform", UNMATCHED_COLUMNS, gaps)

    temporary_excel = excel_path.with_name(f".{excel_path.stem}.tmp.xlsx")
    temporary_csv = csv_path.with_name(f".{csv_path.stem}.tmp.csv")
    try:
        workbook.save(temporary_excel)
        with temporary_csv.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)
        temporary_excel.replace(excel_path)
        temporary_csv.replace(csv_path)
    finally:
        workbook.close()
        temporary_excel.unlink(missing_ok=True)
        temporary_csv.unlink(missing_ok=True)

    by_platforms: dict[str, int] = {}
    for row in rows:
        by_platforms[str(row["platforms"])] = (
            by_platforms.get(str(row["platforms"]), 0) + 1
        )
    return ComparisonResult(
        excel_path=excel_path,
        csv_path=csv_path,
        matches_written=len(rows),
        review_rows=len(review),
        unmatched_rows=len(gaps),
        platforms=tuple(platforms),
        matches_by_platforms=by_platforms,
    )


def _product_from_row(row: Mapping[str, Any]) -> Product:
    """Build a Product from a database row, ignoring bookkeeping columns."""
    values: dict[str, Any] = {}
    for name in Product.__dataclass_fields__:
        value = row.get(name)
        if value is not None:
            values[name] = value
    for numeric in ("mrp", "selling_price", "discount_pct", "rating"):
        if numeric in values:
            values[numeric] = _number(values[numeric])
    for counter in ("rating_count", "review_count"):
        number = _number(values.get(counter))
        values[counter] = int(number) if number is not None else None
    values.setdefault("site", str(row.get("site") or ""))
    values.setdefault("product_id", str(row.get("product_id") or ""))
    values.setdefault("brand", str(row.get("brand") or ""))
    values.setdefault("product_name", str(row.get("product_name") or ""))
    for list_field in (
        "categories", "source_categories", "image_urls", "key_ingredients",
        "key_features", "special_features", "rating_breakdown", "top_reviews",
    ):
        current = values.get(list_field)
        if isinstance(current, str):
            try:
                decoded = json.loads(current)
            except json.JSONDecodeError:
                decoded = []
            values[list_field] = decoded if isinstance(decoded, list) else []
        elif not isinstance(current, list):
            values[list_field] = []
    attributes = values.get("product_attributes")
    if not isinstance(attributes, dict):
        values["product_attributes"] = {}
    return Product(**values)
