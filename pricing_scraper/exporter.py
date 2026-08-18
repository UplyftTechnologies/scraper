"""Structured Excel and CSV export for normalized products."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Callable, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from pricing_scraper.database import (
    DatabaseSyncResult,
    sync_products_to_database,
)
from pricing_scraper.models import Product

OUTPUT_COLUMNS = (
    "site",
    "parent_product_id",
    "product_id",
    "sku",
    "gtin",
    "categories",
    "source_categories",
    "brand",
    "product_name",
    "variant",
    "mrp",
    "selling_price",
    "discount_pct",
    "rating",
    "rating_count",
    "review_count",
    "in_stock",
    "product_url",
    "image_url",
    "image_urls",
    "description",
    "description_html",
    "ingredients",
    "key_ingredients",
    "how_to_use",
    "key_features",
    "special_features",
    "product_attributes",
    "rating_breakdown",
    "top_reviews",
    "scraped_at",
)

IMAGE_COLUMNS = (
    "site",
    "parent_product_id",
    "product_id",
    "sku",
    "variant",
    "image_position",
    "image_url",
)
REVIEW_COLUMNS = (
    "site",
    "parent_product_id",
    "review_id",
    "rating",
    "title",
    "review",
    "reviewer",
    "verified_buyer",
    "created_at",
    "likes",
    "image_urls",
)
EXCEL_MAX_ROWS = 1_048_576


@dataclass(frozen=True, slots=True)
class ExportResult:
    """Paths and counts produced by one export operation."""

    excel_path: Path
    csv_path: Path
    products_written: int
    products_by_site: dict[str, int]
    database_enabled: bool = False
    database_products_written: int = 0
    database_price_points_written: int = 0
    database_error: str = ""


def deduplicate(products: Iterable[Product]) -> list[Product]:
    """Keep the newest SKU observation while merging category memberships."""
    unique: dict[tuple[str, str], Product] = {}
    for product in products:
        key = (product.site.casefold(), product.product_id)
        current = unique.get(key)
        if current is None:
            unique[key] = product
            continue
        newest, older = (
            (product, current)
            if product.scraped_at >= current.scraped_at
            else (current, product)
        )
        attributes = dict(older.product_attributes)
        attributes.update(newest.product_attributes)
        unique[key] = replace(
            newest,
            categories=sorted(
                set(current.categories) | set(product.categories)
            ),
            source_categories=sorted(
                set(current.source_categories)
                | set(product.source_categories)
            ),
            image_urls=list(
                dict.fromkeys([*newest.image_urls, *older.image_urls])
            ),
            key_features=list(
                dict.fromkeys(
                    [*newest.key_features, *older.key_features]
                )
            ),
            special_features=list(
                dict.fromkeys(
                    [*newest.special_features, *older.special_features]
                )
            ),
            product_attributes=attributes,
        )
    return sorted(
        unique.values(),
        key=lambda item: (
            item.site.casefold(),
            item.brand.casefold(),
            item.product_name.casefold(),
            item.variant.casefold(),
        ),
    )


def load_products_csv(path: Path) -> list[Product]:
    """Load normalized products from a previous combined CSV export."""
    if not path.exists():
        return []

    numeric_fields = {
        "mrp": float,
        "selling_price": float,
        "discount_pct": float,
        "rating": float,
        "rating_count": int,
        "review_count": int,
    }
    collection_fields = {
        "categories": [],
        "source_categories": [],
        "image_urls": [],
        "key_ingredients": [],
        "key_features": [],
        "special_features": [],
        "rating_breakdown": [],
        "top_reviews": [],
    }
    products: list[Product] = []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            values: dict[str, object] = {}
            for name in Product.__dataclass_fields__:
                raw = row.get(name, "")
                if name in numeric_fields:
                    try:
                        values[name] = (
                            numeric_fields[name](float(raw))
                            if raw not in ("", None)
                            else None
                        )
                    except (TypeError, ValueError):
                        values[name] = None
                elif name == "in_stock":
                    folded = str(raw or "").casefold()
                    values[name] = (
                        True
                        if folded in {"true", "1"}
                        else False if folded in {"false", "0"} else None
                    )
                elif name == "product_attributes":
                    try:
                        decoded = json.loads(str(raw or ""))
                    except json.JSONDecodeError:
                        decoded = {}
                    values[name] = decoded if isinstance(decoded, dict) else {}
                elif name in collection_fields:
                    try:
                        decoded = json.loads(str(raw or ""))
                    except json.JSONDecodeError:
                        decoded = collection_fields[name]
                    values[name] = (
                        decoded
                        if isinstance(decoded, list)
                        else collection_fields[name]
                    )
                else:
                    values[name] = str(raw or "")
            try:
                products.append(Product(**values))
            except TypeError:
                continue
    return products


def merge_with_existing_sites(
    products: Iterable[Product],
    csv_path: Path | None,
    *,
    replacing_site: str,
) -> list[Product]:
    """Keep prior retailer rows while replacing one site's current snapshot."""
    current = list(products)
    if csv_path is None:
        return deduplicate(current)
    previous = [
        product
        for product in load_products_csv(csv_path.resolve())
        if product.site.casefold() != replacing_site.casefold()
    ]
    return deduplicate([*previous, *current])


def _cell_value(value: object, *, excel: bool) -> object:
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    if excel and isinstance(value, str) and len(value) > 32767:
        return value[:32764] + "..."
    return value


def _row(product: Product, *, excel: bool = False) -> list[object]:
    return [
        _cell_value(getattr(product, column), excel=excel)
        for column in OUTPUT_COLUMNS
    ]


RUPEES = "₹#,##0.00"
PERCENT = '0.00"%"'


class _Widths:
    """Track the widest value seen per column while the rows are written.

    Sizing columns afterwards meant walking every cell in the sheet a second
    time through openpyxl, which for a ten-thousand product catalogue is
    millions of Python-level cell reads and took the better part of an hour.
    The values are already in hand as each row is appended, so their lengths
    are measured then and nothing is re-read.
    """

    __slots__ = ("widest",)

    def __init__(self, headers: Sequence[str]) -> None:
        self.widest = [len(str(header)) for header in headers]

    def observe(self, values: Sequence[object]) -> None:
        widest = self.widest
        for index, value in enumerate(values):
            if value is None or index >= len(widest):
                continue
            length = len(value) if isinstance(value, str) else len(str(value))
            if length > widest[index]:
                widest[index] = length

    def apply(self, sheet: Worksheet) -> None:
        for index, width in enumerate(self.widest, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = min(
                max(width + 2, 10), 60
            )


def _style_header(sheet: Worksheet, widths: _Widths) -> None:
    """Freeze and style the header, then size the columns that were measured."""
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="D81B60")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    widths.apply(sheet)


def _number_format_columns(headers: Sequence[str]) -> dict[int, str]:
    """Which one-based columns need a currency or percentage format."""
    formats: dict[int, str] = {}
    for index, header in enumerate(headers, start=1):
        if header in ("mrp", "selling_price"):
            formats[index] = RUPEES
        elif header == "discount_pct":
            formats[index] = PERCENT
    return formats


def _append_row(
    sheet: Worksheet,
    values: Sequence[object],
    widths: _Widths,
    formats: Mapping[int, str],
    row: int,
) -> None:
    """Append one row, measuring it and formatting only the cells that need it.

    The row number is passed in rather than read back from the sheet.
    ``Worksheet.max_row`` looks like a cheap attribute but rebuilds a set of
    every populated cell's row index on each access, so asking for it once per
    row turns writing a sheet into quadratic work.
    """
    sheet.append(list(values))
    widths.observe(values)
    for column, number_format in formats.items():
        # Three cells per row rather than every cell in the sheet.
        sheet.cell(row=row, column=column).number_format = number_format


def _add_sheet(workbook: Workbook, name: str, products: Iterable[Product]) -> None:
    sheet = workbook.create_sheet(title=name[:31])
    sheet.append(list(OUTPUT_COLUMNS))
    widths = _Widths(OUTPUT_COLUMNS)
    formats = _number_format_columns(OUTPUT_COLUMNS)
    row_number = 1  # the header
    for product in products:
        row_number += 1
        _append_row(sheet, _row(product, excel=True), widths, formats, row_number)
    _style_header(sheet, widths)


def _add_images_sheet(
    workbook: Workbook,
    products: Iterable[Product],
) -> None:
    """Write one row per SKU image so galleries remain structured in Excel."""
    sheets: list[Worksheet] = [workbook.create_sheet(title="images")]
    sheets[0].append(list(IMAGE_COLUMNS))
    sheet = sheets[0]
    widths = [_Widths(IMAGE_COLUMNS)]
    rows_on_sheet = 1  # the header
    seen: set[tuple[str, str, str]] = set()
    for product in products:
        urls = product.image_urls or (
            [product.image_url] if product.image_url else []
        )
        position = 0
        for url in urls:
            text = str(url or "").strip()
            key = (product.site.casefold(), product.product_id, text)
            if not text or key in seen:
                continue
            seen.add(key)
            position += 1
            if rows_on_sheet >= EXCEL_MAX_ROWS:
                sheet = workbook.create_sheet(
                    title=f"images_{len(sheets) + 1}"
                )
                sheet.append(list(IMAGE_COLUMNS))
                sheets.append(sheet)
                widths.append(_Widths(IMAGE_COLUMNS))
                rows_on_sheet = 1
            row = [
                product.site,
                product.parent_product_id,
                product.product_id,
                product.sku,
                product.variant,
                position,
                text,
            ]
            sheet.append(row)
            widths[-1].observe(row)
            rows_on_sheet += 1
    for image_sheet, image_widths in zip(sheets, widths):
        _style_header(image_sheet, image_widths)


def _add_reviews_sheet(
    workbook: Workbook,
    products: Iterable[Product],
) -> None:
    """Write highlighted PDP reviews once per parent/review ID."""
    sheet = workbook.create_sheet(title="reviews")
    sheet.append(list(REVIEW_COLUMNS))
    widths = _Widths(REVIEW_COLUMNS)
    seen: set[tuple[str, str, str]] = set()
    for product in products:
        parent_id = product.parent_product_id or product.product_id
        for review in product.top_reviews:
            if not isinstance(review, dict):
                continue
            review_id = str(review.get("review_id") or "").strip()
            identity = review_id or "|".join(
                str(review.get(field) or "")
                for field in ("created_at", "reviewer", "title", "review")
            )
            key = (product.site.casefold(), parent_id, identity)
            if key in seen:
                continue
            seen.add(key)
            row = [
                product.site,
                parent_id,
                review_id,
                review.get("rating"),
                review.get("title"),
                review.get("review"),
                review.get("reviewer"),
                review.get("verified_buyer"),
                review.get("created_at"),
                review.get("likes"),
                _cell_value(review.get("images", []), excel=True),
            ]
            sheet.append(row)
            widths.observe(row)
    _style_header(sheet, widths)


def export_products(
    products: Iterable[Product],
    output_path: Path,
    csv_path: Path | None = None,
    *,
    sync_database: bool = False,
    write_excel: bool = True,
    status_callback: Callable[[str], None] | None = None,
) -> ExportResult:
    """Persist products to the database, Excel, and combined CSV.

    ``write_excel=False`` writes the CSV and skips the workbook. Building the
    workbook holds every cell in memory - around 500 MB for a ten-thousand
    product catalogue - which a small hosted instance cannot afford, and on a
    container with a temporary disk the file would not survive a restart
    anyway.
    """
    normalized = deduplicate(products)
    if not normalized:
        raise ValueError("No products are available to export.")

    excel_path = output_path.resolve()
    combined_csv_path = (
        csv_path.resolve()
        if csv_path
        else excel_path.with_name(f"{excel_path.stem}_combined.csv")
    )
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    combined_csv_path.parent.mkdir(parents=True, exist_ok=True)

    grouped: dict[str, list[Product]] = {}
    for product in normalized:
        grouped.setdefault(product.site.casefold(), []).append(product)

    if status_callback is not None:
        status_callback(
            f"Writing {len(normalized):,} products to "
            + ("Excel and CSV..." if write_excel else "CSV...")
        )

    temporary_csv = combined_csv_path.with_name(
        f".{combined_csv_path.stem}.tmp.csv"
    )
    try:
        # The CSV is streamed row by row, so it costs almost no memory even for
        # a large catalogue. Write it first: if the workbook cannot be built,
        # the run still leaves a complete export behind.
        with temporary_csv.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow(OUTPUT_COLUMNS)
            writer.writerows(_row(product) for product in normalized)
        temporary_csv.replace(combined_csv_path)
    finally:
        temporary_csv.unlink(missing_ok=True)

    if write_excel:
        workbook = Workbook()
        workbook.remove(workbook.active)
        _add_sheet(workbook, "combined", normalized)
        for site, site_products in sorted(grouped.items()):
            _add_sheet(workbook, site, site_products)
        _add_images_sheet(workbook, normalized)
        _add_reviews_sheet(workbook, normalized)

        temporary_excel = excel_path.with_name(f".{excel_path.stem}.tmp.xlsx")
        try:
            workbook.save(temporary_excel)
            temporary_excel.replace(excel_path)
        finally:
            workbook.close()
            temporary_excel.unlink(missing_ok=True)

    # The database is synchronized only once the local files are on disk. With
    # DATABASE_SYNC_REQUIRED set, a failed write still raises and still stops
    # the run being called complete - but a transient Supabase timeout can no
    # longer throw away the export of a collection that took hours.
    if status_callback is not None and sync_database:
        status_callback(
            f"Synchronizing {len(normalized):,} products with Supabase..."
        )
    database = (
        sync_products_to_database(normalized)
        if sync_database
        else DatabaseSyncResult(enabled=False)
    )

    return ExportResult(
        excel_path=excel_path,
        csv_path=combined_csv_path,
        products_written=len(normalized),
        products_by_site={
            site: len(site_products) for site, site_products in grouped.items()
        },
        database_enabled=database.enabled,
        database_products_written=database.products_written,
        database_price_points_written=database.price_points_written,
        database_error=database.error,
    )
